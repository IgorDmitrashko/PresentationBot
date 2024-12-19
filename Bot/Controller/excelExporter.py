import openpyxl
from bs4 import CData
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def __init__(self, filename):
        """
        Инициализация класса для экспорта данных в Excel.

        :param filename: Имя файла Excel для сохранения данных.
        """
        self.filename = filename

    def export(self, data):
        """
        Экспортирует данные из списка объектов в Excel-файл.

        :param data: Список объектов, каждый из которых содержит атрибуты name, price и image.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        headers = ["Name", "Price", "Image"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        for row_num, item in enumerate(data, start=2):
            sheet.cell(row=row_num, column=1, value=item.name)
            sheet.cell(row=row_num, column=2, value=item.price)

            image_cell = sheet.cell(row=row_num, column=3, value="View Image")
            image_cell.hyperlink = item.image  # Ссылка на файл изображения
            image_cell.font = Font(color="0000FF", underline="single")  # Стиль гиперссылки

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True

        workbook.save(self.filename)
        print(f"Данные успешно сохранены в файл {self.filename}")


